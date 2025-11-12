import os
import shutil
from datetime import datetime
from typing import Dict, Optional
from fastapi import UploadFile, Form
from pypdf import PdfWriter
import pikepdf
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from docx2pdf import convert
from pdf2docx import Converter
from typing import List, Dict, Optional
import tabula
import pandas as pd
import pymupdf as fitz
import uuid
import shutil
from datetime import datetime
from typing import Dict
from fastapi import UploadFile, HTTPException
from pathlib import Path
import subprocess
import math
from typing import Dict, Optional
from fastapi import UploadFile
import asyncio
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware


UPLOAD_DIR = "uploads"
OUTPUT_DIR = "outputs"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

def get_temp_path(filename: str, prefix: str, upload_dir: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(upload_dir, f"{prefix}_{timestamp}_{filename}")

def get_output_path(prefix: str, output_dir: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(output_dir, f"{prefix}_{timestamp}.pdf")
#working  merge
def merge_two_pdfs(
    file_1: UploadFile,
    file_2: UploadFile,
    upload_dir: str,
    output_dir: str,
    library: str = "pypdf"
) -> Dict[str, str]:
    """
    Merges exactly two PDFs using the specified library.

    Supported libraries:
        pypdf, PyPDF2, pdfrw, fitz (PyMuPDF), PyPDF4, PyPDF3, pdfplumber, pdfminer.six

    Returns dict with the filename of the merged PDF.
    """
    library = library.lower().strip()

    # -------------------------------
    # 1. Validate file extensions
    # -------------------------------
    for f in (file_1, file_2):
        if not f.filename.lower().endswith('.pdf'):
            raise ValueError(f"File '{f.filename}' is not a PDF.")

    # -------------------------------
    # 2. Save uploaded files temporarily
    # -------------------------------
    temp_paths = []
    try:
        temp_path1 = get_temp_path(file_1.filename, "merge1", upload_dir)
        temp_path2 = get_temp_path(file_2.filename, "merge2", upload_dir)

        with open(temp_path1, "wb") as buf:
            shutil.copyfileobj(file_1.file, buf)
        with open(temp_path2, "wb") as buf:
            shutil.copyfileobj(file_2.file, buf)

        temp_paths = [temp_path1, temp_path2]

        # -------------------------------
        # 3. Merge using selected library
        # -------------------------------
        output_path = get_output_path("merged", output_dir)

        if library == "pypdf":
            from pypdf import PdfWriter
            writer = PdfWriter()
            writer.append(temp_path1)
            writer.append(temp_path2)
            with open(output_path, "wb") as f:
                writer.write(f)

        elif library == "pypdf2":
            from PyPDF2 import PdfWriter, PdfReader
            writer = PdfWriter()
            for path in temp_paths:
                reader = PdfReader(path)
                for page in reader.pages:
                    writer.add_page(page)
            with open(output_path, "wb") as f:
                writer.write(f)

        elif library == "pdfrw":
            from pdfrw import PdfWriter, PdfReader
            writer = PdfWriter()
            for path in temp_paths:
                reader = PdfReader(path)
                writer.addpages(reader.pages)
            writer.write(output_path)

        elif library in ("pymupdf", "fitz"):
            import fitz  # PyMuPDF
            doc = fitz.Document()
            for path in temp_paths:
                src = fitz.open(path)
                doc.insert_pdf(src)
                src.close()
            doc.save(output_path)
            doc.close()

        elif library == "pypdf4":
            from PyPDF4 import PdfFileWriter, PdfFileReader
            writer = PdfFileWriter()
            for path in temp_paths:
                reader = PdfFileReader(path)
                for page_num in range(reader.getNumPages()):
                    writer.addPage(reader.getPage(page_num))
            with open(output_path, "wb") as f:
                writer.write(f)

        elif library == "pypdf3":
            from PyPDF3 import PdfFileWriter, PdfFileReader
            writer = PdfFileWriter()
            for path in temp_paths:
                reader = PdfFileReader(path)
                for page_num in range(reader.getNumPages()):
                    writer.addPage(reader.getPage(page_num))
            with open(output_path, "wb") as f:
                writer.write(f)

        elif library == "pdfplumber":
            import pdfplumber
            from pypdf import PdfWriter  # fallback writer
            writer = PdfWriter()
            for path in temp_paths:
                with pdfplumber.open(path) as pdf:
                    for page in pdf.pages:
                        # pdfplumber doesn't support direct merge, so use pypdf as backend
                        writer.append(path, pages=(page.page_number - 1, page.page_number))
            with open(output_path, "wb") as f:
                writer.write(f)
        else:
            raise ValueError(f"Unsupported library: {library}. "
                             f"Choose from: pypdf, PyPDF2, pdfrw, fitz, PyPDF4, PyPDF3, pdfplumber")

        return {"filename": os.path.basename(output_path)}

    finally:
        for p in temp_paths:
            if os.path.exists(p):
                try:
                    os.remove(p)
                except OSError:
                    pass  # ignore if already deleted



# new compress



def _calculate_target_size(
    original_bytes: int,
    target_size: Optional[str] = None,
    percent: Optional[float] = None,
) -> int:
    """
    Priority:
      1. target_size → "500KB" or "2MB"
      2. percent → 70 (means 70% of original)
      3. Default → 50% of original
    """
    if target_size:
        s = target_size.strip().upper()
        if not s.endswith(("KB", "MB")):
            raise ValueError("target_size must end with 'KB' or 'MB'")
        try:
            val = float(s[:-2])
        except ValueError:
            raise ValueError("Invalid number in target_size")
        return int(val * 1024) if s.endswith("KB") else int(val * 1024 * 1024)

    if percent is not None:
        if not (1 <= percent <= 99):
            raise ValueError("compression_percent must be between 1 and 99")
        return int(original_bytes * (percent / 100))

    return int(original_bytes * 0.5)          # <-- your default 50%


# ----------------------------------------------------------------------
# 2. Image-binary-search helper (your code, cleaned up)
# ----------------------------------------------------------------------
def _compress_with_images(input_path: str, output_path: str, target_bytes: int) -> bool:
    """
    Returns True if the result is ≤ target_bytes * 1.1
    (10 % tolerance – enough for most PDFs)
    """
    import fitz

    doc = fitz.open(input_path)
    low, high = 0.3, 1.0
    best_zoom = 1.0

    for _ in range(15):
        zoom = (low + high) / 2
        size = _render_and_measure(doc, output_path + ".tmp", zoom)
        if size <= target_bytes:
            best_zoom = zoom
            low = zoom
        else:
            high = zoom

    # final render
    _render_and_measure(doc, output_path, best_zoom, final=True)
    doc.close()

    if os.path.exists(output_path + ".tmp"):
        os.remove(output_path + ".tmp")

    achieved = os.path.getsize(output_path)
    return achieved <= target_bytes * 1.1


def _render_and_measure(doc, temp_path: str, zoom: float, final: bool = False) -> int:
    import fitz
    new_doc = fitz.open()
    for page in doc:
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        fmt = "jpeg" if final else "png"
        img_data = pix.tobytes(fmt)
        new_page = new_doc.new_page(width=page.rect.width, height=page.rect.height)
        new_page.insert_image(page.rect, stream=img_data)
    new_doc.save(temp_path, garbage=4, deflate=True, clean=True)
    new_doc.close()
    return os.path.getsize(temp_path)


# ----------------------------------------------------------------------
# 3. Library-specific compressors (unchanged, except they receive target_bytes)
# ----------------------------------------------------------------------
def _compress_with_fitz(input_path: str, output_path: str, target_bytes: int) -> None:
    import fitz
    doc = fitz.open(input_path)
    low, high = 0.05, 1.0
    best_zoom = 1.0
    for _ in range(16):
        zoom = (low + high) / 2
        size = _render_and_measure(doc, output_path + ".tmp", zoom)
        if size <= target_bytes:
            best_zoom = zoom
            low = zoom
        else:
            high = zoom
    _render_and_measure(doc, output_path, best_zoom, final=True)
    doc.close()
    if os.path.exists(output_path + ".tmp"):
        os.remove(output_path + ".tmp")


def _compress_with_ghostscript(input_path: str, output_path: str, target_bytes: int) -> None:
    if shutil.which("gs") is None:
        raise RuntimeError("Ghostscript not found – brew install ghostscript")
    orig = os.path.getsize(input_path)
    ratio = target_bytes / orig
    dpi = max(50, min(300, int(math.sqrt(ratio / 0.0007))))
    cmd = [
        "gs", "-q", "-dNOPAUSE", "-dBATCH", "-dSAFER",
        "-sDEVICE=pdfwrite", "-dCompatibilityLevel=1.4",
        "-dPDFSETTINGS=/ebook",
        f"-dColorImageResolution={dpi}",
        f"-dGrayImageResolution={dpi}",
        f"-dMonoImageResolution={dpi}",
        f"-sOutputFile={output_path}", input_path
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f"Ghostscript error: {r.stderr.strip()}")


def _compress_with_pikepdf(input_path: str, output_path: str) -> None:
    import pikepdf
    with pikepdf.open(input_path) as pdf:
        pdf.save(output_path, compress_streams=True, linearize=True)


def _compress_with_qpdf(input_path: str, output_path: str) -> None:
    cmd = ["qpdf", "--linearize", "--compress-streams=y", input_path, output_path]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f"qpdf error: {r.stderr.strip()}")


def _compress_with_pdfminer_reportlab(input_path: str, output_path: str, target_bytes: int) -> None:
    from pdfminer.high_level import extract_text
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    text = extract_text(input_path)
    c = canvas.Canvas(output_path, pagesize=letter)
    w, h = letter
    y = h - 50
    for line in text.splitlines():
        if y < 50:
            c.showPage()
            y = h - 50
        c.drawString(50, y, line[:120])
        y -= 14
    c.save()


def _compress_with_pdfrw(input_path: str, output_path: str) -> None:
    from pdfrw import PdfReader, PdfWriter
    r = PdfReader(input_path)
    w = PdfWriter()
    w.addpages(r.pages)
    w.write(output_path)


def _compress_with_pypdf2(input_path: str, output_path: str) -> None:
    from PyPDF2 import PdfReader, PdfWriter
    r = PdfReader(input_path)
    w = PdfWriter()
    for p in r.pages:
        w.add_page(p)
    with open(output_path, "wb") as f:
        w.write(f)


def _compress_with_pypdf(input_path: str, output_path: str) -> None:
    from pypdf import PdfWriter
    w = PdfWriter()
    w.append(input_path)
    with open(output_path, "wb") as f:
        w.write(f)


# ----------------------------------------------------------------------
# 4. COMMON COMPRESS – default image path + fallback
# ----------------------------------------------------------------------
def _common_compress(
    input_path: str,
    final_output: str,
    target_bytes: int,
    library: str,
) -> str:
    """
    Returns the path of the final file.
    """
    temp_output = final_output + ".tmp"

    # ---------- 1. Try the image-binary-search (default) ----------
    try:
        if _compress_with_images(input_path, temp_output, target_bytes):
            os.rename(temp_output, final_output)
            return final_output
    except Exception as e:
        # image path failed – continue to library fallback
        if os.path.exists(temp_output):
            os.remove(temp_output)

    # ---------- 2. Library-specific fallback ----------
    if library in ("pymupdf", "fitz"):
        _compress_with_fitz(input_path, final_output, target_bytes)
    elif library == "ghostscript":
        _compress_with_ghostscript(input_path, final_output, target_bytes)
    elif library == "pikepdf":
        _compress_with_pikepdf(input_path, final_output)
    elif library == "qpdf":
        _compress_with_qpdf(input_path, final_output)
    elif library == "pdfminer.six":
        _compress_with_pdfminer_reportlab(input_path, final_output, target_bytes)
    elif library == "pdfrw":
        _compress_with_pdfrw(input_path, final_output)
    elif library == "pypdf2":
        _compress_with_pypdf2(input_path, final_output)
    elif library == "pypdf":
        _compress_with_pypdf(input_path, final_output)
    else:
        raise ValueError(f"Unsupported library: {library}")

    return final_output


# ----------------------------------------------------------------------
# 5. Public API – compress_pdf()
# ----------------------------------------------------------------------
def compress_pdf(
    file_1: UploadFile,
    upload_dir: str,
    output_dir: str,
    target_size: Optional[str] = None,
    compression_percent: Optional[float] = None,
    library: str = "fitz",               # default = image-binary-search
) -> Dict[str, str]:
    library = library.lower().strip()

    if not file_1.filename.lower().endswith('.pdf'):
        raise ValueError("File must be a PDF.")

    os.makedirs(upload_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    temp_input   = os.path.join(upload_dir, f"in_{file_1.filename}")
    final_output = os.path.join(output_dir, f"compressed_{file_1.filename}")

    try:
        # ---- save uploaded file ----
        with open(temp_input, "wb") as f:
            shutil.copyfileobj(file_1.file, f)

        orig_bytes   = os.path.getsize(temp_input)
        target_bytes = _calculate_target_size(orig_bytes, target_size, compression_percent)

        # ---- compress (default image path → library fallback) ----
        compressed_path = _common_compress(temp_input, final_output, target_bytes, library)

        final_size = os.path.getsize(compressed_path)
        accuracy   = min(999.9, round(final_size / target_bytes * 100, 1))

        return {
            "filename": os.path.basename(compressed_path),
            "original_size_kb": round(orig_bytes / 1024, 1),
            "compressed_size_kb": round(final_size / 1024, 1),
            "target_accuracy": f"{accuracy}%",
            "target_bytes": target_bytes,
        }

    finally:
        # clean only temporary files
        for p in (temp_input, final_output + ".tmp"):
            if os.path.exists(p):
                try:
                    os.remove(p)
                except:
                    pass

#close compress

def cleanup_temp_files(temp_paths: List[str]) -> None:
    for path in temp_paths:
        if os.path.exists(path):
            os.remove(path)

def get_temp_path(filename: str, prefix: str, upload_dir: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(upload_dir, f"{prefix}_{timestamp}_{filename}")

def get_output_path(prefix: str, output_dir: str, extension: str = ".pdf") -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{timestamp}{extension}"
    return os.path.join(output_dir, filename)
# not used 
def merge_pdfs(file_1: UploadFile, file_2: UploadFile, upload_dir: str, output_dir: str) -> Dict[str, str]:
    files = [file_1, file_2]
    if len(files) < 2:
        raise ValueError("At least two PDFs required.")
    temp_paths = []
    try:
        for file in files:
            if not file.filename.lower().endswith('.pdf'):
                raise ValueError(f"File {file.filename} is not a PDF.")
            temp_path = get_temp_path(file.filename, "merge_temp", upload_dir)
            with open(temp_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            temp_paths.append(temp_path)
        
        writer = PdfWriter()
        for temp_path in temp_paths:
            writer.append(temp_path)
        
        output_path = get_output_path("merged", output_dir)
        with open(output_path, "wb") as f:
            writer.write(f)
        writer.close()
        
        return {"filename": os.path.basename(output_path)}
    
    except Exception as e:
        raise e
    
    finally:
        cleanup_temp_files(temp_paths)

def compress_pdf_old(file_1: UploadFile, upload_dir: str, output_dir: str) -> Dict[str, str]:
    if not file_1.filename.lower().endswith('.pdf'):
        raise ValueError("File is not a PDF.")
    temp_path = get_temp_path(file_1.filename, "compress_temp", upload_dir)
    try:
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file_1.file, buffer)
        
        output_path = get_output_path("compressed", output_dir)
        with pikepdf.open(temp_path) as pdf:
            pdf.save(output_path, optimize_images=True, compression_level=9)
        
        return {"filename": os.path.basename(output_path)}
    
    except Exception as e:
        raise e
    
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)
# working fing upto this functionality

def excel_to_pdf(
    file_excel: UploadFile,
    upload_dir: str,
    output_dir: str,
    sheet_name: Optional[str] = None,     # e.g. "Sales"
    sheet_index: Optional[int] = None     # e.g. 0, 1, 2
) -> Dict[str, str]:
    """
    Convert Excel to PDF with flexible sheet selection:
    - sheet_name="Sales" → by name
    - sheet_index=1 → by index (0-based)
    - both None → ALL sheets (multi-page PDF)
    """
    from openpyxl import load_workbook
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    import os
    import shutil
    from fastapi import HTTPException

    if not file_excel.filename.lower().endswith(('.xlsx', '.xls')):
        raise ValueError("File must be Excel (.xlsx or .xls).")

    temp_path = get_temp_path(file_excel.filename, "excel_temp", upload_dir)
    output_path = get_output_path("excel_to_pdf", output_dir)

    try:
        with open(temp_path, "wb") as f:
            shutil.copyfileobj(file_excel.file, f)

        wb = load_workbook(temp_path, data_only=True)
        doc = SimpleDocTemplate(output_path, pagesize=letter, topMargin=40)
        elements = []
        styles = getSampleStyleSheet()

        # === Determine sheets to convert ===
        sheets = []

        if sheet_name and sheet_index is not None:
            raise ValueError("Use either sheet_name OR sheet_index, not both.")

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
            sheets = [wb[sheet_name]]

        elif sheet_index is not None:
            if not (0 <= sheet_index < len(wb.worksheets)):
                raise ValueError(f"sheet_index {sheet_index} out of range. Valid: 0–{len(wb.worksheets)-1}")
            sheets = [wb.worksheets[sheet_index]]

        else:
            sheets = wb.worksheets  # ALL sheets

        # === Process Each Sheet ===
        for ws in sheets:
            title = Paragraph(f"<b>{ws.title}</b>", styles["Title"])
            elements.append(title)
            elements.append(Spacer(1, 12))

            data = []
            for row in ws.iter_rows(values_only=True):
                safe_row = ["" if c is None else str(c).strip() for c in row]
                data.append(safe_row)

            data = [row for row in data if any(cell for cell in row)]
            if not data:
                elements.append(Paragraph("No data in this sheet.", styles["Normal"]))
                elements.append(PageBreak())
                continue

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(table)
            elements.append(PageBreak())

        if not elements:
            raise ValueError("No data found in selected sheet(s).")

        doc.build(elements)

        mode = "all_sheets" if not sheet_name and sheet_index is None else "by_name" if sheet_name else "by_index"
        return {
            "filename": os.path.basename(output_path),
            "sheets_converted": len(sheets),
            "sheet_names": ", ".join(ws.title for ws in sheets),
            "mode": mode
        }

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error converting Excel to PDF: {str(e)}")

    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

# above code is working fineee
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
LIBREOFFICE_TIMEOUT = 60


def docx_to_pdfs( file_excel: UploadFile,
    upload_dir: str,
    output_dir: str, output_pdf=None):
    """
    Convert a .docx file to .pdf
    """
    if not os.path.exists(input_docx):
        print(f"Error: File {input_docx} not found!")
        return False

    if output_pdf is None:
        output_pdf = input_docx.replace(".docx", ".pdf")

    try:
        convert(input_docx, output_pdf)
        print(f"Successfully converted: {input_docx} → {output_pdf}")
        return True
    except Exception as e:
        print(f"Conversion failed: {e}")
        return False


async def word_to_pdf(file_doc: UploadFile, upload_dir: str, output_dir: str) -> Dict[str, str]:
    """
    Convert .docx → PDF using docx2pdf (pure pip package).
    Works on macOS, Windows, Linux.
    """
    original_name = Path(file_doc.filename).name

    # Only .docx supported
    if not original_name.lower().endswith('.docx'):
        raise HTTPException(400, "Only .docx files are supported with docx2pdf")

    # Paths
    input_path = os.path.join(upload_dir, f"docx_{uuid.uuid4().hex}.docx")
    final_pdf = os.path.join(output_dir, f"docx_{uuid.uuid4().hex}.pdf")
    cleanup = [input_path]

    try:
        # Save uploaded .docx
        os.makedirs(upload_dir, exist_ok=True)
        with open(input_path, "wb") as f:
            shutil.copyfileobj(file_doc.file, f)

        # Convert: docx2pdf writes directly to final_pdf
        os.makedirs(output_dir, exist_ok=True)
        docx_convert(input_path, final_pdf)

        if not os.path.exists(final_pdf):
            raise HTTPException(500, "PDF was not generated by docx2pdf")

        return {
            "filename": Path(final_pdf).name,
            "original_format": "docx",
            "method": "docx2pdf"
        }

    except Exception as e:
        raise HTTPException(500, f"Conversion failed: {str(e)}")
    finally:
        # Clean up input file
        for p in cleanup:
            if os.path.exists(p):
                try:
                    os.remove(p)
                except:
                    pass


def pdf_to_word(file_pdf: UploadFile, upload_dir: str, output_dir: str) -> Dict[str, str]:
    if not file_pdf.filename.lower().endswith('.pdf'):
        raise ValueError("File must be .pdf.")
    temp_path = get_temp_path(file_pdf.filename, "pdf_to_word_temp", upload_dir)
    try:
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file_pdf.file, buffer)
        
        output_path = get_output_path("pdf_to_word", output_dir, ".docx")
        cv = Converter(temp_path)
        cv.convert(output_path, start=0, end=None)
        cv.close()
        
        return {"filename": os.path.basename(output_path)}
    
    except Exception as e:
        raise e
    
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

def pdf_to_excel(file_pdf: UploadFile, upload_dir: str, output_dir: str, pages: Optional[str] = "all") -> Dict[str, str]:
    if not file_pdf.filename.lower().endswith('.pdf'):
        raise ValueError("File must be .pdf.")
    temp_path = get_temp_path(file_pdf.filename, "pdf_to_excel_temp", upload_dir)
    try:
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file_pdf.file, buffer)
        
        dfs = tabula.read_pdf(temp_path, pages=pages, multiple_tables=True)
        if not dfs:
            raise ValueError("No tables found in PDF.")
        output_path = get_output_path("pdf_to_excel", output_dir, ".xlsx")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for i, df in enumerate(dfs):
                df.to_excel(writer, sheet_name=f"Table_{i+1}", index=False)
        
        return {"filename": os.path.basename(output_path)}
    
    except Exception as e:
        raise e
    
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)
# new pdf functions
